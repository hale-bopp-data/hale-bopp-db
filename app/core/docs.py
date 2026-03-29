"""hb docs — Documentation generators for the Data Dictionary.

Generates: Mermaid ER, DBML, Markdown reference, Excel (6 sheets), HTML.

Phase 1: PBI #546.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from app.core.compile import DataDictionary, EntityDef, resolve_type


# ---------------------------------------------------------------------------
# Mermaid ER diagram (from dictionary, not live DB)
# ---------------------------------------------------------------------------

def generate_mermaid(dictionary: DataDictionary, schema_filter: str | None = None) -> str:
    """Generate Mermaid erDiagram from the data dictionary."""
    entities = dictionary.entities
    if schema_filter:
        entities = [e for e in entities if e.schema_name == schema_filter]

    entity_names = {e.name for e in entities}
    lines = ["erDiagram"]

    # Relationships
    for rel in dictionary.relationships:
        from_e = rel.get("from", "")
        to_e = rel.get("to", "")
        if schema_filter and (from_e not in entity_names or to_e not in entity_names):
            continue
        card = rel.get("cardinality", "many-to-one")
        mermaid_card = _cardinality_to_mermaid(card)
        label = rel.get("from_column", "")
        lines.append(f'    {to_e} {mermaid_card} {from_e} : "{label}"')

    # Entities with columns
    for entity in entities:
        lines.append(f"    {entity.name} {{")
        pk_cols = set()
        if entity.pk:
            pk_cols = set(entity.pk.get("columns", []))
        for col in entity.columns:
            pg_type = resolve_type(col.type, dictionary.type_map, "pg")
            # Sanitize type for Mermaid (no parentheses)
            safe_type = pg_type.replace("(", "_").replace(")", "").replace(",", "_")
            markers = []
            if col.name in pk_cols or col.pk:
                markers.append("PK")
            if col.fk:
                markers.append("FK")
            marker_str = " " + ",".join(markers) if markers else ""
            comment = ""
            if col.description:
                safe_desc = col.description.replace('"', "'")[:60]
                comment = f' "{safe_desc}"'
            lines.append(f"        {safe_type} {col.name}{marker_str}{comment}")
        lines.append("    }")

    return "\n".join(lines)


def _cardinality_to_mermaid(card: str) -> str:
    mapping = {
        "one-to-one": "||--||",
        "one-to-many": "||--o{",
        "many-to-one": "||--o{",
        "many-to-many": "}o--o{",
    }
    return mapping.get(card, "||--o{")


# ---------------------------------------------------------------------------
# DBML export (for dbdiagram.io)
# ---------------------------------------------------------------------------

def generate_dbml(dictionary: DataDictionary, schema_filter: str | None = None) -> str:
    """Generate DBML from the data dictionary."""
    entities = dictionary.entities
    if schema_filter:
        entities = [e for e in entities if e.schema_name == schema_filter]

    lines: list[str] = []

    for entity in entities:
        pk_cols = set()
        if entity.pk:
            pk_cols = set(entity.pk.get("columns", []))

        lines.append(f"Table {entity.schema_name}.{entity.name} {{")

        for col in entity.columns:
            pg_type = resolve_type(col.type, dictionary.type_map, "pg")
            settings: list[str] = []
            if col.name in pk_cols or col.pk:
                settings.append("pk")
            if not col.nullable:
                settings.append("not null")
            if col.default is not None:
                settings.append(f"default: {col.default}")
            if col.description:
                safe = col.description.replace("'", "\\'")
                settings.append(f"note: '{safe}'")

            settings_str = f" [{', '.join(settings)}]" if settings else ""
            lines.append(f"  {col.name} {pg_type}{settings_str}")

        # Indexes
        if entity.indexes:
            lines.append("")
            lines.append("  indexes {")
            for idx in entity.indexes:
                cols = ", ".join(idx.columns)
                opts = []
                if idx.unique:
                    opts.append("unique")
                opts.append(f"name: '{idx.name}'")
                lines.append(f"    ({cols}) [{', '.join(opts)}]")
            lines.append("  }")

        lines.append("}")
        lines.append("")

    # Relationships as Ref
    for rel in dictionary.relationships:
        from_e = rel.get("from", "")
        to_e = rel.get("to", "")
        from_col = rel.get("from_column", "")
        to_col = rel.get("to_column", "")
        if schema_filter:
            entity_names = {e.name for e in entities}
            if from_e not in entity_names or to_e not in entity_names:
                continue
        # Find schemas
        schema_map = {e.name: e.schema_name for e in dictionary.entities}
        from_schema = schema_map.get(from_e, "public")
        to_schema = schema_map.get(to_e, "public")
        lines.append(f"Ref: {from_schema}.{from_e}.{from_col} > {to_schema}.{to_e}.{to_col}")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Markdown reference
# ---------------------------------------------------------------------------

def generate_markdown(dictionary: DataDictionary, schema_filter: str | None = None) -> str:
    """Generate Markdown documentation from the data dictionary."""
    entities = dictionary.entities
    if schema_filter:
        entities = [e for e in entities if e.schema_name == schema_filter]

    lines: list[str] = []
    lines.append("# Data Dictionary Reference")
    lines.append("")

    # TOC by schema
    schemas: dict[str, list[EntityDef]] = {}
    for e in entities:
        schemas.setdefault(e.schema_name, []).append(e)

    lines.append("## Table of Contents")
    lines.append("")
    for schema_name in sorted(schemas):
        lines.append(f"### {schema_name}")
        for e in schemas[schema_name]:
            anchor = f"{e.schema_name}-{e.name}".replace("_", "-")
            lines.append(f"- [{e.schema_name}.{e.name}](#{anchor}) — {e.description or ''}")
        lines.append("")

    lines.append("---")
    lines.append("")

    # Entity details
    for schema_name in sorted(schemas):
        for entity in schemas[schema_name]:
            anchor = f"{entity.schema_name}-{entity.name}".replace("_", "-")
            lines.append(f"## {entity.schema_name}.{entity.name}")
            lines.append("")
            if entity.description:
                lines.append(f"**Descrizione**: {entity.description}")
            if entity.description_nonna:
                lines.append(f"")
                lines.append(f"> *Per la nonna*: {entity.description_nonna}")
            lines.append("")
            lines.append(f"**Tipo**: {entity.type} | **Schema**: {entity.schema_name} | **Multi-tenant**: {entity.multi_tenant}")
            lines.append("")

            # Columns table
            lines.append("| Colonna | Tipo | Nullable | Default | PK | FK | PII | Descrizione |")
            lines.append("|---------|------|----------|---------|----|----|-----|-------------|")

            pk_cols = set()
            if entity.pk:
                pk_cols = set(entity.pk.get("columns", []))

            for col in entity.columns:
                pg_type = resolve_type(col.type, dictionary.type_map, "pg")
                nullable = "NULL" if col.nullable else "NOT NULL"
                default = col.default or ""
                is_pk = "PK" if (col.name in pk_cols or col.pk) else ""
                is_fk = col.fk or ""
                pii = "PII" if col.pii else ""
                desc = col.description or ""
                if col.description_nonna:
                    desc += f" (*{col.description_nonna}*)"
                lines.append(f"| {col.name} | {pg_type} | {nullable} | {default} | {is_pk} | {is_fk} | {pii} | {desc} |")

            lines.append("")

            # Indexes
            if entity.indexes:
                lines.append("**Indici**:")
                for idx in entity.indexes:
                    unique = "UNIQUE " if idx.unique else ""
                    lines.append(f"- `{idx.name}`: {unique}({', '.join(idx.columns)})")
                lines.append("")

            # Security
            if entity.security:
                sec = entity.security
                rls = sec.get("rls", False)
                pii_cols = sec.get("pii_columns", [])
                masking = sec.get("masking", {})
                if rls or pii_cols or masking:
                    lines.append("**Sicurezza**:")
                    if rls:
                        lines.append(f"- RLS: {sec.get('rls_policy', 'da configurare')}")
                    if pii_cols:
                        lines.append(f"- PII: {', '.join(pii_cols)}")
                    if masking:
                        lines.append(f"- Masking: {', '.join(f'{k}={v}' for k, v in masking.items())}")
                    lines.append("")

            lines.append("---")
            lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Excel export (6 sheets)
# ---------------------------------------------------------------------------

def generate_excel(dictionary: DataDictionary, output_path: str | Path) -> Path:
    """Generate Excel file with 6 sheets from the data dictionary."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    def _write_header(ws, headers: list[str]) -> None:
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # Sheet 1: Entities
    ws = wb.active
    ws.title = "Entità"
    _write_header(ws, ["Schema", "Nome", "Tipo", "Multi-tenant", "Descrizione", "Descrizione Nonna", "Colonne", "Indici"])
    for i, e in enumerate(dictionary.entities, 2):
        ws.cell(row=i, column=1, value=e.schema_name)
        ws.cell(row=i, column=2, value=e.name)
        ws.cell(row=i, column=3, value=e.type)
        ws.cell(row=i, column=4, value=str(e.multi_tenant))
        ws.cell(row=i, column=5, value=e.description or "")
        ws.cell(row=i, column=6, value=e.description_nonna or "")
        ws.cell(row=i, column=7, value=len(e.columns))
        ws.cell(row=i, column=8, value=len(e.indexes))

    # Sheet 2: Columns
    ws2 = wb.create_sheet("Colonne")
    _write_header(ws2, ["Schema", "Tabella", "Colonna", "Tipo Logico", "Tipo PG", "Nullable", "Default", "PK", "FK", "PII", "Check", "Descrizione"])
    row = 2
    for e in dictionary.entities:
        pk_cols = set(e.pk.get("columns", [])) if e.pk else set()
        for col in e.columns:
            pg_type = resolve_type(col.type, dictionary.type_map, "pg")
            ws2.cell(row=row, column=1, value=e.schema_name)
            ws2.cell(row=row, column=2, value=e.name)
            ws2.cell(row=row, column=3, value=col.name)
            ws2.cell(row=row, column=4, value=col.type)
            ws2.cell(row=row, column=5, value=pg_type)
            ws2.cell(row=row, column=6, value="YES" if col.nullable else "NO")
            ws2.cell(row=row, column=7, value=col.default or "")
            ws2.cell(row=row, column=8, value="PK" if col.name in pk_cols or col.pk else "")
            ws2.cell(row=row, column=9, value=col.fk or "")
            ws2.cell(row=row, column=10, value="PII" if col.pii else "")
            ws2.cell(row=row, column=11, value=col.check or "")
            ws2.cell(row=row, column=12, value=col.description or "")
            row += 1

    # Sheet 3: Relationships
    ws3 = wb.create_sheet("Relazioni")
    _write_header(ws3, ["Da", "Colonna Da", "A", "Colonna A", "Cardinalità"])
    for i, rel in enumerate(dictionary.relationships, 2):
        ws3.cell(row=i, column=1, value=rel.get("from", ""))
        ws3.cell(row=i, column=2, value=rel.get("from_column", ""))
        ws3.cell(row=i, column=3, value=rel.get("to", ""))
        ws3.cell(row=i, column=4, value=rel.get("to_column", ""))
        ws3.cell(row=i, column=5, value=rel.get("cardinality", ""))

    # Sheet 4: Security
    ws4 = wb.create_sheet("Sicurezza")
    _write_header(ws4, ["Schema", "Tabella", "RLS", "Policy RLS", "Colonne PII", "Masking"])
    row = 2
    for e in dictionary.entities:
        sec = e.security
        if not sec:
            continue
        rls = sec.get("rls", False)
        pii = sec.get("pii_columns", [])
        masking = sec.get("masking", {})
        if rls or pii or masking:
            ws4.cell(row=row, column=1, value=e.schema_name)
            ws4.cell(row=row, column=2, value=e.name)
            ws4.cell(row=row, column=3, value="YES" if rls else "NO")
            ws4.cell(row=row, column=4, value=sec.get("rls_policy", ""))
            ws4.cell(row=row, column=5, value=", ".join(pii))
            ws4.cell(row=row, column=6, value=", ".join(f"{k}={v}" for k, v in masking.items()))
            row += 1

    # Sheet 5: Type Map
    ws5 = wb.create_sheet("Type Map")
    engines = ["pg", "mssql", "ora", "syn", "rs"]
    _write_header(ws5, ["Tipo Logico"] + [e.upper() for e in engines])
    for i, (logical, mapping) in enumerate(dictionary.type_map.items(), 2):
        ws5.cell(row=i, column=1, value=logical)
        for j, eng in enumerate(engines, 2):
            ws5.cell(row=i, column=j, value=mapping.get(eng, ""))

    # Sheet 6: Name Mapping
    ws6 = wb.create_sheet("Mapping Nomi")
    _write_header(ws6, ["Concetto", "Nome Canonico", "Alias Vietati"])
    for i, nm in enumerate(dictionary.name_mapping, 2):
        ws6.cell(row=i, column=1, value=nm.get("concept", ""))
        ws6.cell(row=i, column=2, value=nm.get("canonical", ""))
        ws6.cell(row=i, column=3, value=", ".join(nm.get("aliases_forbidden", [])))

    # Auto-width columns for all sheets
    for ws_item in wb.worksheets:
        for col_cells in ws_item.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws_item.column_dimensions[col_letter].width = min(max_len + 2, 50)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return out


# ---------------------------------------------------------------------------
# HTML static page
# ---------------------------------------------------------------------------

def generate_html(dictionary: DataDictionary, schema_filter: str | None = None) -> str:
    """Generate a single-page HTML with Mermaid ER and entity tables."""
    entities = dictionary.entities
    if schema_filter:
        entities = [e for e in entities if e.schema_name == schema_filter]

    mermaid_code = generate_mermaid(dictionary, schema_filter=schema_filter)

    # Group by schema
    schemas: dict[str, list[EntityDef]] = {}
    for e in entities:
        schemas.setdefault(e.schema_name, []).append(e)

    # Build entity cards
    entity_cards = []
    for schema_name in sorted(schemas):
        for entity in schemas[schema_name]:
            pk_cols = set(entity.pk.get("columns", [])) if entity.pk else set()
            rows = []
            for col in entity.columns:
                pg_type = resolve_type(col.type, dictionary.type_map, "pg")
                badges = []
                if col.name in pk_cols or col.pk:
                    badges.append('<span class="badge pk">PK</span>')
                if col.fk:
                    badges.append(f'<span class="badge fk">FK→{col.fk}</span>')
                if col.pii:
                    badges.append('<span class="badge pii">PII</span>')
                nullable = "NULL" if col.nullable else "NOT NULL"
                default = col.default or ""
                desc = col.description or ""
                desc_nonna = f'<br><em class="nonna">{col.description_nonna}</em>' if col.description_nonna else ""
                rows.append(
                    f"<tr><td><code>{col.name}</code></td><td>{pg_type}</td>"
                    f"<td>{nullable}</td><td>{default}</td>"
                    f"<td>{' '.join(badges)}</td><td>{desc}{desc_nonna}</td></tr>"
                )

            desc_section = ""
            if entity.description:
                desc_section += f"<p>{entity.description}</p>"
            if entity.description_nonna:
                desc_section += f'<p class="nonna">Per la nonna: {entity.description_nonna}</p>'

            card = f"""
    <div class="entity-card" data-schema="{entity.schema_name}" data-name="{entity.name}" data-search="{entity.schema_name} {entity.name} {entity.description or ''} {entity.description_nonna or ''}">
      <h3 id="{entity.schema_name}-{entity.name}">{entity.schema_name}.{entity.name}
        <span class="entity-type">{entity.type}</span>
      </h3>
      {desc_section}
      <table>
        <thead><tr><th>Colonna</th><th>Tipo</th><th>Nullable</th><th>Default</th><th>Info</th><th>Descrizione</th></tr></thead>
        <tbody>{''.join(rows)}</tbody>
      </table>
    </div>"""
            entity_cards.append(card)

    # TOC
    toc_items = []
    for schema_name in sorted(schemas):
        toc_items.append(f'<h4>{schema_name}</h4><ul>')
        for e in schemas[schema_name]:
            toc_items.append(f'<li><a href="#{e.schema_name}-{e.name}">{e.name}</a> <small>{e.type}</small></li>')
        toc_items.append("</ul>")

    return _HTML_TEMPLATE.format(
        title="Data Dictionary — hale-bopp-db",
        entity_count=len(entities),
        schema_count=len(schemas),
        toc="".join(toc_items),
        mermaid=mermaid_code,
        cards="".join(entity_cards),
    )


_HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="it">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{title}</title>
  <script src="https://cdn.jsdelivr.net/npm/mermaid@11/dist/mermaid.min.js"></script>
  <style>
    :root {{ --bg: #f8f9fa; --card: #fff; --border: #dee2e6; --accent: #2f5496; --text: #212529; }}
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', system-ui, sans-serif;
           background: var(--bg); color: var(--text); line-height: 1.6; padding: 2rem; }}
    h1 {{ color: var(--accent); margin-bottom: 0.5rem; }}
    h3 {{ color: var(--accent); margin-bottom: 0.5rem; border-bottom: 2px solid var(--accent); padding-bottom: 0.3rem; }}
    .stats {{ color: #6c757d; margin-bottom: 1.5rem; }}
    .search {{ width: 100%; padding: 0.7rem 1rem; font-size: 1rem; border: 2px solid var(--border);
               border-radius: 8px; margin-bottom: 1.5rem; }}
    .search:focus {{ outline: none; border-color: var(--accent); }}
    .layout {{ display: grid; grid-template-columns: 250px 1fr; gap: 2rem; }}
    .sidebar {{ position: sticky; top: 1rem; max-height: calc(100vh - 3rem); overflow-y: auto; }}
    .sidebar h4 {{ color: var(--accent); margin-top: 1rem; }}
    .sidebar ul {{ list-style: none; padding-left: 0.5rem; }}
    .sidebar li {{ margin: 0.2rem 0; }}
    .sidebar a {{ text-decoration: none; color: var(--text); }}
    .sidebar a:hover {{ color: var(--accent); }}
    .sidebar small {{ color: #6c757d; font-size: 0.75rem; }}
    .entity-card {{ background: var(--card); border: 1px solid var(--border); border-radius: 8px;
                    padding: 1.5rem; margin-bottom: 1.5rem; }}
    .entity-type {{ font-size: 0.75rem; background: var(--accent); color: #fff; padding: 0.15rem 0.5rem;
                    border-radius: 4px; vertical-align: middle; font-weight: normal; }}
    table {{ width: 100%; border-collapse: collapse; margin-top: 0.5rem; font-size: 0.9rem; }}
    th {{ background: var(--accent); color: #fff; padding: 0.5rem; text-align: left; }}
    td {{ padding: 0.4rem 0.5rem; border-bottom: 1px solid var(--border); }}
    tr:hover {{ background: #f0f4ff; }}
    code {{ background: #e9ecef; padding: 0.1rem 0.3rem; border-radius: 3px; font-size: 0.85rem; }}
    .badge {{ font-size: 0.7rem; padding: 0.1rem 0.4rem; border-radius: 3px; margin-right: 0.2rem; }}
    .badge.pk {{ background: #2f5496; color: #fff; }}
    .badge.fk {{ background: #6c757d; color: #fff; }}
    .badge.pii {{ background: #dc3545; color: #fff; }}
    .nonna {{ color: #6c757d; font-style: italic; }}
    .er-diagram {{ background: var(--card); border: 1px solid var(--border); border-radius: 8px;
                   padding: 1.5rem; margin-bottom: 1.5rem; overflow-x: auto; }}
    .hidden {{ display: none; }}
    @media (max-width: 768px) {{ .layout {{ grid-template-columns: 1fr; }} .sidebar {{ position: static; }} }}
  </style>
</head>
<body>
  <h1>{title}</h1>
  <p class="stats">{entity_count} entità | {schema_count} schema | Generato da hale-bopp-db</p>
  <input type="text" class="search" id="search" placeholder="Cerca entità, colonne, descrizioni..." autofocus>

  <div class="er-diagram">
    <h3>ER Diagram</h3>
    <pre class="mermaid">
{mermaid}
    </pre>
  </div>

  <div class="layout">
    <nav class="sidebar">{toc}</nav>
    <main id="entities">{cards}</main>
  </div>

  <script>
    mermaid.initialize({{ startOnLoad: true, theme: 'default' }});
    document.getElementById('search').addEventListener('input', function(e) {{
      const q = e.target.value.toLowerCase();
      document.querySelectorAll('.entity-card').forEach(card => {{
        const text = card.getAttribute('data-search').toLowerCase();
        card.classList.toggle('hidden', q && !text.includes(q));
      }});
    }});
  </script>
</body>
</html>"""
