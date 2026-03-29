"""Tests for hb docs — Documentation generators."""

from __future__ import annotations

from pathlib import Path

import pytest

from app.core.compile import DataDictionary, load_dictionary
from app.core.docs import (
    generate_dbml,
    generate_excel,
    generate_html,
    generate_markdown,
    generate_mermaid,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

def _make_dict(entities: list[dict] | None = None, **overrides) -> DataDictionary:
    base = {
        "type_map": {
            "uuid": {"pg": "TEXT"},
            "string": {"pg": "TEXT"},
            "string(n)": {"pg": "VARCHAR({n})"},
            "integer": {"pg": "INTEGER"},
            "auto": {"pg": "BIGSERIAL"},
            "boolean": {"pg": "BOOLEAN"},
            "timestamp": {"pg": "TIMESTAMPTZ"},
            "json": {"pg": "JSONB"},
            "text_large": {"pg": "TEXT"},
            "long": {"pg": "BIGINT"},
            "date": {"pg": "DATE"},
            "decimal(p,s)": {"pg": "NUMERIC({p},{s})"},
        },
        "default_map": {"now()": {"pg": "NOW()"}, "true": {"pg": "TRUE"}},
        "name_mapping": [
            {"concept": "Tenant", "canonical": "tenant", "aliases_forbidden": ["COD_CLI"]},
        ],
        "entities": entities or [TENANT, USER],
        "relationships": [
            {"from": "user", "from_column": "tenant_id", "to": "tenant", "to_column": "tenant_id", "cardinality": "many-to-one"},
        ],
    }
    base.update(overrides)
    return DataDictionary.model_validate(base)


TENANT = {
    "name": "tenant",
    "schema": "platform",
    "type": "DIM",
    "description": "Root entity — every row belongs to a tenant.",
    "description_nonna": "L'azienda cliente.",
    "multi_tenant": "root",
    "pk": {"columns": ["id"], "type": "auto"},
    "security": {},
    "columns": [
        {"name": "id", "type": "auto", "nullable": False, "description": "Surrogate key"},
        {"name": "tenant_id", "type": "string(50)", "nullable": False, "description": "Business key"},
        {"name": "tenant_name", "type": "string(255)", "nullable": False},
        {"name": "created_at", "type": "timestamp", "nullable": False, "default": "now()"},
        {"name": "created_by", "type": "string", "nullable": False},
        {"name": "updated_at", "type": "timestamp", "nullable": False},
        {"name": "updated_by", "type": "string", "nullable": False},
    ],
}

USER = {
    "name": "user",
    "schema": "platform",
    "type": "DIM",
    "description": "Portal user.",
    "description_nonna": "Una persona che usa il portale.",
    "multi_tenant": True,
    "pk": {"columns": ["id"], "type": "auto"},
    "security": {"rls": True, "pii_columns": ["email"], "masking": {"email": "partial"}},
    "columns": [
        {"name": "id", "type": "auto", "nullable": False},
        {"name": "tenant_id", "type": "string(50)", "nullable": False, "fk": "tenant.tenant_id"},
        {"name": "email", "type": "string(320)", "nullable": False, "pii": True, "description": "User email"},
        {"name": "created_at", "type": "timestamp", "nullable": False},
        {"name": "created_by", "type": "string", "nullable": False},
        {"name": "updated_at", "type": "timestamp", "nullable": False},
        {"name": "updated_by", "type": "string", "nullable": False},
    ],
}


# ---------------------------------------------------------------------------
# Mermaid
# ---------------------------------------------------------------------------

class TestMermaid:
    def test_starts_with_erdiagram(self):
        dd = _make_dict()
        result = generate_mermaid(dd)
        assert result.startswith("erDiagram")

    def test_contains_entities(self):
        dd = _make_dict()
        result = generate_mermaid(dd)
        assert "tenant {" in result
        assert "user {" in result

    def test_contains_relationships(self):
        dd = _make_dict()
        result = generate_mermaid(dd)
        assert "tenant" in result
        assert "user" in result
        assert "tenant_id" in result

    def test_contains_pk_marker(self):
        dd = _make_dict()
        result = generate_mermaid(dd)
        assert "PK" in result

    def test_contains_fk_marker(self):
        dd = _make_dict()
        result = generate_mermaid(dd)
        assert "FK" in result

    def test_schema_filter(self):
        dd = _make_dict()
        result = generate_mermaid(dd, schema_filter="nonexistent")
        assert "tenant" not in result or result.count("{") == 0


# ---------------------------------------------------------------------------
# DBML
# ---------------------------------------------------------------------------

class TestDbml:
    def test_contains_tables(self):
        dd = _make_dict()
        result = generate_dbml(dd)
        assert "Table platform.tenant {" in result
        assert "Table platform.user {" in result

    def test_contains_pk(self):
        dd = _make_dict()
        result = generate_dbml(dd)
        assert "[pk" in result

    def test_contains_refs(self):
        dd = _make_dict()
        result = generate_dbml(dd)
        assert "Ref:" in result
        assert "platform.user.tenant_id" in result

    def test_contains_indexes(self):
        entity_with_idx = {
            **TENANT,
            "indexes": [{"name": "ux_tenant_id", "columns": ["tenant_id"], "unique": True}],
        }
        dd = _make_dict([entity_with_idx])
        result = generate_dbml(dd)
        assert "indexes {" in result
        assert "ux_tenant_id" in result


# ---------------------------------------------------------------------------
# Markdown
# ---------------------------------------------------------------------------

class TestMarkdown:
    def test_title(self):
        dd = _make_dict()
        result = generate_markdown(dd)
        assert "# Data Dictionary Reference" in result

    def test_contains_toc(self):
        dd = _make_dict()
        result = generate_markdown(dd)
        assert "## Table of Contents" in result

    def test_contains_entity_header(self):
        dd = _make_dict()
        result = generate_markdown(dd)
        assert "## platform.tenant" in result
        assert "## platform.user" in result

    def test_contains_columns_table(self):
        dd = _make_dict()
        result = generate_markdown(dd)
        assert "| Colonna |" in result
        assert "| tenant_id |" in result

    def test_contains_description_nonna(self):
        dd = _make_dict()
        result = generate_markdown(dd)
        assert "Per la nonna" in result
        assert "L'azienda cliente" in result

    def test_contains_security_info(self):
        dd = _make_dict()
        result = generate_markdown(dd)
        assert "RLS" in result
        assert "PII" in result


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

class TestExcel:
    def test_creates_file(self, tmp_path: Path):
        dd = _make_dict()
        path = generate_excel(dd, output_path=tmp_path / "dict.xlsx")
        assert path.exists()
        assert path.suffix == ".xlsx"

    def test_has_6_sheets(self, tmp_path: Path):
        from openpyxl import load_workbook
        dd = _make_dict()
        path = generate_excel(dd, output_path=tmp_path / "dict.xlsx")
        wb = load_workbook(str(path))
        assert len(wb.sheetnames) == 6
        assert "Entità" in wb.sheetnames
        assert "Colonne" in wb.sheetnames
        assert "Relazioni" in wb.sheetnames
        assert "Sicurezza" in wb.sheetnames
        assert "Type Map" in wb.sheetnames
        assert "Mapping Nomi" in wb.sheetnames

    def test_entities_sheet_data(self, tmp_path: Path):
        from openpyxl import load_workbook
        dd = _make_dict()
        path = generate_excel(dd, output_path=tmp_path / "dict.xlsx")
        wb = load_workbook(str(path))
        ws = wb["Entità"]
        # Header + 2 entities = 3 rows
        assert ws.max_row == 3
        assert ws.cell(row=2, column=2).value == "tenant"

    def test_columns_sheet_data(self, tmp_path: Path):
        from openpyxl import load_workbook
        dd = _make_dict()
        path = generate_excel(dd, output_path=tmp_path / "dict.xlsx")
        wb = load_workbook(str(path))
        ws = wb["Colonne"]
        # Header + 14 columns (7+7)
        assert ws.max_row == 15


# ---------------------------------------------------------------------------
# HTML
# ---------------------------------------------------------------------------

class TestHtml:
    def test_valid_html(self):
        dd = _make_dict()
        result = generate_html(dd)
        assert "<!DOCTYPE html>" in result
        assert "</html>" in result

    def test_contains_mermaid(self):
        dd = _make_dict()
        result = generate_html(dd)
        assert "mermaid" in result
        assert "erDiagram" in result

    def test_contains_entity_cards(self):
        dd = _make_dict()
        result = generate_html(dd)
        assert "platform.tenant" in result
        assert "platform.user" in result

    def test_contains_search(self):
        dd = _make_dict()
        result = generate_html(dd)
        assert 'id="search"' in result
        assert "Cerca" in result

    def test_contains_nonna(self):
        dd = _make_dict()
        result = generate_html(dd)
        assert "L&#x27;azienda cliente" in result or "L'azienda cliente" in result

    def test_contains_pii_badge(self):
        dd = _make_dict()
        result = generate_html(dd)
        assert "PII" in result

    def test_contains_toc(self):
        dd = _make_dict()
        result = generate_html(dd)
        assert "sidebar" in result


# ---------------------------------------------------------------------------
# Real dictionary
# ---------------------------------------------------------------------------

REAL_DICT_PATH = Path("C:/EW/easyway/wiki/guides/db-data-dictionary.json")


@pytest.mark.skipif(not REAL_DICT_PATH.exists(), reason="Real dictionary not found")
class TestRealDictionary:
    def test_mermaid(self):
        dd = load_dictionary(REAL_DICT_PATH)
        result = generate_mermaid(dd)
        assert "erDiagram" in result
        assert "tenant" in result

    def test_dbml(self):
        dd = load_dictionary(REAL_DICT_PATH)
        result = generate_dbml(dd)
        assert "Table platform.tenant" in result
        assert "Ref:" in result

    def test_markdown(self):
        dd = load_dictionary(REAL_DICT_PATH)
        result = generate_markdown(dd)
        assert "# Data Dictionary Reference" in result
        assert result.count("##") > 20  # 20 entities + TOC

    def test_excel(self, tmp_path: Path):
        dd = load_dictionary(REAL_DICT_PATH)
        path = generate_excel(dd, output_path=tmp_path / "real.xlsx")
        assert path.exists()
        from openpyxl import load_workbook
        wb = load_workbook(str(path))
        ws = wb["Entità"]
        assert ws.max_row >= 21  # header + 20 entities

    def test_html(self):
        dd = load_dictionary(REAL_DICT_PATH)
        result = generate_html(dd)
        assert "<!DOCTYPE html>" in result
        assert result.count("entity-card") >= 20

    def test_full_generate(self, tmp_path: Path):
        dd = load_dictionary(REAL_DICT_PATH)
        out = tmp_path / "docs"
        out.mkdir()
        (out / "index.html").write_text(generate_html(dd), encoding="utf-8")
        (out / "data-dictionary.md").write_text(generate_markdown(dd), encoding="utf-8")
        (out / "er-diagram.mmd").write_text(generate_mermaid(dd), encoding="utf-8")
        (out / "schema.dbml").write_text(generate_dbml(dd), encoding="utf-8")
        generate_excel(dd, output_path=out / "dictionary.xlsx")
        assert (out / "index.html").exists()
        assert (out / "data-dictionary.md").exists()
        assert (out / "er-diagram.mmd").exists()
        assert (out / "schema.dbml").exists()
        assert (out / "dictionary.xlsx").exists()
